# -*- coding: utf-8 -*-
"""Страницы выгрузки: ассортименты → рецепты / SKU / прайс-лист.

Формула прайса (решение от 04.08.2026):
  цена = total_sku_cost (без НДС) × 1.22 (НДС 22%) × коэффициент
  Базовый  ×2.4
  Партнёр  ×2.1
Округление до целых рублей.
"""
import re
from decimal import Decimal, ROUND_HALF_UP

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, Http404
from django.shortcuts import render, get_object_or_404

from .models import (
    Assortment, AssortmentProduct, AssortmentRecipe,
    Client, RawMaterialWithPrice, Recipe, RecipeClient, RecipeCost, SkuCost,
)


def _nat_key(s):
    """Натуральная сортировка артикулов: 21 < 100, 4101/3 < 4101/12."""
    return [int(t) if t.isdigit() else t.lower()
            for t in re.split(r'(\d+)', s or '')]

VAT = Decimal('1.22')
TIERS = (
    ('base', 'Базовый', Decimal('2.4')),
    ('partner', 'Партнёр', Decimal('2.1')),
)

TITLES = {
    'STM': 'СТМ',
    'OLD_TEA': 'Старый ассортимент',
    'NEW_TEA': 'Новый ассортимент',
}


def _rub(value):
    return int(value.quantize(Decimal('1'), rounding=ROUND_HALF_UP))


def _assortment(code):
    if code not in TITLES:
        raise Http404
    return get_object_or_404(Assortment, code=code)


def _client(request, code):
    """Выбранный СТМ-клиент (?client=<uuid>), только для STM."""
    cid = request.GET.get('client')
    if code != 'STM' or not cid:
        return None
    return get_object_or_404(Client, pk=cid)


def _ctx(request, code, section=''):
    a = _assortment(code)
    client = _client(request, code)
    ctx = {
        'a': a,
        'code': code,
        'title': TITLES[code],
        'section': section,
        'client': client,
        'client_qs': f'?client={client.id}' if client else '',
    }
    if code == 'STM':
        ctx['clients'] = Client.objects.order_by('name')
    return ctx


def _fmt_package(size, unit):
    if size is None:
        return ''
    s = f'{size.normalize():f}'.rstrip('0').rstrip('.')
    return f'{s} {unit or ""}'.strip()


@login_required
def home(request):
    counts = {}
    for ap in AssortmentProduct.objects.select_related('assortment'):
        counts[ap.assortment.code] = counts.get(ap.assortment.code, 0) + 1
    cards = [
        {'code': c, 'title': TITLES[c], 'skus': counts.get(c, 0)}
        for c in ('STM', 'OLD_TEA', 'NEW_TEA')
    ]
    return render(request, 'catalog/home.html', {'cards': cards})


@login_required
def hub(request, code):
    ctx = _ctx(request, code)
    return render(request, 'catalog/hub.html', ctx)


@login_required
def recipes(request, code):
    ctx = _ctx(request, code, 'Рецепты')
    if code == 'STM' and ctx['client']:
        pairs = (RecipeClient.objects.filter(client=ctx['client'])
                 .select_related('recipe'))
    else:
        pairs = (AssortmentRecipe.objects.filter(assortment=ctx['a'])
                 .select_related('recipe'))
    ctx['items'] = sorted((p.recipe for p in pairs),
                          key=lambda r: _nat_key(r.article))
    return render(request, 'catalog/recipes.html', ctx)


def _sku_qs(ctx):
    qs = (AssortmentProduct.objects.filter(assortment=ctx['a'])
          .select_related('product', 'product__type')
          .order_by('product__article'))
    if ctx['client']:
        qs = qs.filter(client=ctx['client'])
    return qs


@login_required
def skus(request, code):
    ctx = _ctx(request, code, 'SKU')
    items = []
    for ap in _sku_qs(ctx):
        p = ap.product
        items.append({
            'article': p.article,
            'name': p.name,
            'package': _fmt_package(p.package_size, p.package_unit),
            'type': p.type.name if p.type else '',
            'active': bool(p.is_active),
        })
    items.sort(key=lambda i: _nat_key(i['article']))
    ctx['items'] = items
    return render(request, 'catalog/skus.html', ctx)


def _price_rows(ctx):
    aps = [ap for ap in _sku_qs(ctx) if ap.product.is_active]
    ids = [ap.product.id for ap in aps]
    costs = {c.product_id: c.total_sku_cost
             for c in SkuCost.objects.filter(product_id__in=ids)}
    rows = []
    for ap in aps:
        p = ap.product
        cost = costs.get(p.id)
        if not cost:
            continue
        cost_vat = cost * VAT
        rows.append({
            'article': p.article,
            'name': p.name,
            'package': _fmt_package(p.package_size, p.package_unit),
            'prices': {key: _rub(cost_vat * k) for key, _, k in TIERS},
        })
    rows.sort(key=lambda r: _nat_key(r['article']))
    return rows


@login_required
def price(request, code):
    ctx = _ctx(request, code, 'Прайс-лист')
    ctx['rows'] = _price_rows(ctx)
    ctx['tiers'] = TIERS
    return render(request, 'catalog/price.html', ctx)


@login_required
def recipe_detail(request, pk):
    r = get_object_or_404(Recipe, pk=pk)
    ings = list(r.ingredients.select_related('material', 'sub_recipe'))

    mat_ids = [i.material_id for i in ings if i.material_id]
    sub_ids = [i.sub_recipe_id for i in ings if i.sub_recipe_id]
    prices = {m.id: m.current_price
              for m in RawMaterialWithPrice.objects.filter(id__in=mat_ids)}
    costs = {c.recipe_id: c
             for c in RecipeCost.objects.filter(recipe_id__in=sub_ids + [r.id])}

    rows, total_qty = [], Decimal('0')
    for i in ings:
        qty = i.quantity or Decimal('0')
        total_qty += qty
        if i.material_id:
            price = prices.get(i.material_id)
            rows.append({
                'article': i.material.article, 'name': i.material.name,
                'is_sub': False, 'sub_id': None,
                'qty': qty, 'unit': i.unit,
                'price': price,
                'line': (qty * price) if price is not None else None,
            })
        elif i.sub_recipe_id:
            sub_cost = costs.get(i.sub_recipe_id)
            price = sub_cost.cost_per_kg if sub_cost else None
            rows.append({
                'article': i.sub_recipe.article, 'name': i.sub_recipe.name,
                'is_sub': True, 'sub_id': i.sub_recipe_id,
                'qty': qty, 'unit': i.unit,
                'price': price,
                'line': (qty * price) if price is not None else None,
            })
    for row in rows:
        row['grams'] = int(row['qty'] * 1000)
        row['share'] = (row['qty'] / total_qty * 100) if total_qty else None
    rows.sort(key=lambda x: x['qty'], reverse=True)

    my_cost = costs.get(r.id)
    back_code = request.GET.get('a') if request.GET.get('a') in TITLES else None
    back_client = request.GET.get('client', '')
    return render(request, 'catalog/recipe.html', {
        'r': r,
        'rows': rows,
        'total_qty': total_qty,
        'total_grams': int(total_qty * 1000),
        'cost': my_cost,
        'back_code': back_code,
        'back_title': TITLES.get(back_code, ''),
        'back_qs': f'?client={back_client}' if back_client else '',
    })


@login_required
def price_xlsx(request, code):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from urllib.parse import quote
    import datetime

    ctx = _ctx(request, code, 'Прайс-лист')
    rows = _price_rows(ctx)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Прайс-лист'

    subtitle = TITLES[code] + (f' · {ctx["client"].name}' if ctx['client'] else '')
    today = datetime.date.today().strftime('%d.%m.%Y')

    dark = '494E31'
    ws.merge_cells('A1:E1')
    ws['A1'] = f'ПЧК · Прайс-лист — {subtitle} — {today} (цены с НДС 22%)'
    ws['A1'].font = Font(bold=True, size=12, color=dark)

    headers = ['Артикул', 'Название', 'Фасовка'] + [t[1] + ', ₽' for t in TIERS]
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor=dark)
        c.border = border
        c.alignment = Alignment(horizontal='center')

    r = 4
    for row in rows:
        ws.cell(row=r, column=1, value=row['article']).border = border
        ws.cell(row=r, column=2, value=row['name']).border = border
        ws.cell(row=r, column=3, value=row['package']).border = border
        for i, (key, _, _k) in enumerate(TIERS):
            c = ws.cell(row=r, column=4 + i, value=row['prices'][key])
            c.border = border
            c.number_format = '# ##0'
        r += 1

    widths = [16, 52, 10, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A4'

    name = f'Прайс_{subtitle.replace(" · ", "_").replace(" ", "_")}_{today}.xlsx'
    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(name)}"
    wb.save(resp)
    return resp
